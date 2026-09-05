# -*- coding: utf-8 -*-
"""生成：大模型供应商与旗舰模型评分参考表 (.xlsx)"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "大模型供应商评分参考.xlsx")

HEAD = ["供应商", "国家/地区", "代表旗舰模型", "开源", "综合评分", "推理能力", "编程能力",
        "中文能力", "多模态", "API定价档位", "一句话点评"]

ROWS = [
    ["OpenAI", "美国", "GPT-4o / GPT-4.1 / o3 系列", "否", 9.6, 9.7, 9.6, 9.2, 9.6, "高", "旗舰标杆，生态、工具链与 Agent 最全"],
    ["Anthropic", "美国", "Claude Opus 4 / Sonnet 4", "否", 9.5, 9.5, 9.8, 8.8, 9.3, "高", "编程与长文写作口碑最佳"],
    ["Google", "美国", "Gemini 2.5 Pro / Flash", "部分", 9.3, 9.4, 9.3, 9.0, 9.7, "偏高", "百万级上下文 + 原生多模态"],
    ["DeepSeek", "中国", "DeepSeek-V3.1 / R1", "是", 9.2, 9.4, 9.1, 9.5, 7.0, "低", "开源性价比之王，推理逼近第一梯队"],
    ["阿里通义", "中国", "Qwen3-235B-A22B / Qwen2.5", "是", 8.9, 8.9, 8.8, 9.4, 8.5, "低", "开源生态完善，中文与 Agent 调用强"],
    ["xAI", "美国", "Grok 3 / Grok 4", "部分", 8.9, 9.0, 8.7, 7.8, 8.0, "高", "推理强，多模态与开源布局激进"],
    ["月之暗面", "中国", "Kimi K2 / k1.5", "是", 8.8, 8.8, 9.0, 9.0, 6.5, "中", "开源 Agent/编程表现出众"],
    ["字节豆包", "中国", "Doubao-Seed-1.6 / 1.5-Pro", "否", 8.7, 8.6, 8.6, 9.3, 8.6, "低", "中文交互体验好，价格极低、C 端量大"],
    ["智谱AI", "中国", "GLM-4.6 / GLM-4.5", "部分", 8.6, 8.5, 8.4, 9.2, 8.3, "低", "全面均衡，Agent + 端侧布局深"],
    ["百度", "中国", "ERNIE-4.5 / X1", "是", 8.5, 8.3, 8.1, 9.0, 8.4, "低", "中文理解强，4.5 起转向开源"],
    ["Meta", "美国", "Llama 4 (Maverick/Behemoth)", "是", 8.4, 8.3, 8.4, 7.6, 8.6, "免费", "开源生态源头，社区权重与微调最多"],
    ["腾讯混元", "中国", "Hunyuan-T1 / Turbo", "部分", 8.3, 8.3, 8.0, 8.9, 7.8, "低", "深绑腾讯系应用场景"],
    ["MiniMax", "中国", "MiniMax-M2", "是", 8.3, 8.4, 8.5, 8.4, 7.5, "中", "长上下文 + 音乐多模态特色"],
    ["Mistral", "法国", "Mistral Large 3 / Medium 3", "部分", 8.0, 7.9, 7.9, 6.8, 7.9, "中", "欧洲代表，模型小而精"],
]

NOTE = [
    "口径与说明（重要）",
    "1. 时间：模型与能力评分以 2025 年中公开信息为基准（知识快照）；本表生成于 2026-09-06。厂商迭代极快，2026 年新旗舰未收录，请按需更新后使用。",
    "2. 依据：综合参考 LMArena 主观榜、各类公开榜单（MMLU/GPQA/Bench 等）、开发者实测口碑，属『社区综合参考值』，非任何官方成绩，满分 10 分。",
    "3. 各分项：推理/编程/中文/多模态为该供应商旗舰代表模型的量级估计（多模态一列对纯文本推理模型偏低属正常）。",
    "4. 定价档位：指旗舰 API 的每百万 token 输入价量级 —— 低≈¥1~几元级（如 DeepSeek/豆包），高≈$2.5~15 级（OpenAI/Anthropic）。开源或免费列标注『免费』。",
    "5. 开源列：『是』= 权重开放；『部分』= 部分型号或限权重（如 Gemini/Grok 部分开源）；『否』= 闭源 API 为主。",
    "6. 免责：数字为主观参考，选型请以官方最新 benchmark 与自身业务实测为准。",
]

def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    for c in ws[1]:
        c.fill, c.font = fill, font
        c.alignment = Alignment(horizontal="center", vertical="center")

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "模型评分"
    ws.append(HEAD)
    for r in ROWS:
        ws.append(r)

    widths = [12, 10, 32, 8, 10, 10, 10, 10, 10, 12, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    style_header(ws)
    thin = Side(style="thin", color="D9D9D9")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEAD)):
        for c in row:
            c.border = Border(bottom=thin)
            c.font = Font(name="微软雅黑", size=10)
            if c.column in (5, 6, 7, 8, 9, 10):
                c.alignment = Alignment(horizontal="center", vertical="center")
    for i in range(2, ws.max_row + 1):
        score = ws.cell(row=i, column=5).value
        c = ws.cell(row=i, column=5)
        c.font = Font(name="微软雅黑", size=11, bold=True)
        if score >= 9.0:
            c.fill = PatternFill("solid", fgColor="C6EFCE")
        elif score >= 8.5:
            c.fill = PatternFill("solid", fgColor="FFEB9C")
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("口径说明")
    ws2.column_dimensions["A"].width = 110
    for i, t in enumerate(NOTE, 1):
        ws2.cell(row=i, column=1, value=t).font = Font(name="微软雅黑", size=10, bold=(i == 1))
    wb.save(OUT)
    print("OK:", OUT, "rows=", len(ROWS))

if __name__ == "__main__":
    main()
